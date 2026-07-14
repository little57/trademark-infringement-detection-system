# -*- coding: utf-8 -*-
"""
中国传媒大学 侵权商品检测系统 - GUI 前端
"""
import os
import pathlib
import sys
import threading
import traceback
import webbrowser
import tkinter as tk
from tkinter import messagebox, ttk

BASE = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

from backend.detector import Detector, REPORTS_DIR, SCREENSHOTS, AI_VERIFIER_AVAILABLE

# AI验证报告单独存放目录（与reports同级）
AI_REPORTS_DIR = REPORTS_DIR.parent / "ai_reports"

try:
    from backend.ai_verifier import batch_verify
    AI_BATCH_AVAILABLE = True
except ImportError:
    AI_BATCH_AVAILABLE = False



class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("中国传媒大学 - 侵权商品检测系统")
        self.root.geometry("1100x700")
        self.root.minsize(900, 550)
        self.root.configure(bg="#f5f7fa")

        self.results = []
        self._last_excel = ""
        self._detector = None
        self._running = False
        self._build()

    def _build(self):
        r = self.root
        primary = "#1a3a5c"
        success = "#27ae60"
        danger = "#e74c3c"

        # ========== 顶部标题栏 ==========
        header = tk.Frame(r, bg=primary, height=55)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="CUC", font=("微软雅黑", 16, "bold"), bg=primary, fg="white").pack(
            side="left", padx=(16, 6)
        )
        tk.Label(header, text="侵权商品检测系统", font=("微软雅黑", 12, "bold"), bg=primary, fg="white").pack(
            side="left"
        )

        self.slb = tk.Label(
            header, text="就绪", font=("微软雅黑", 9, "bold"), bg="#95a5a6", fg="white", padx=14, pady=3
        )
        self.slb.pack(side="right", padx=14)

        # ========== 主区域 ==========
        main = tk.Frame(r, bg="#f5f7fa")
        main.pack(fill="both", expand=True, padx=10, pady=(8, 4))

        # ========== 搜索页数输入框 ==========
        search_frame = tk.Frame(main, bg="white", highlightbackground="#dde", highlightthickness=1, padx=10, pady=8)
        search_frame.pack(fill="x", pady=(0, 6))

        tk.Label(search_frame, text="📄 搜索页数：", font=("微软雅黑", 10), bg="white", fg="#2c3e50").pack(side="left")

        self.pages_var = tk.StringVar(value="10")
        self.pages_entry = tk.Entry(
            search_frame,
            textvariable=self.pages_var,
            font=("微软雅黑", 10),
            width=8,
            bd=1,
            relief="solid",
        )
        self.pages_entry.pack(side="left", padx=(6, 10), ipady=3)

        # 开始检测按钮
        self.btn_start = tk.Button(
            search_frame,
            text="🚀 开始检测",
            font=("微软雅黑", 10, "bold"),
            bg=primary,
            fg="white",
            padx=18,
            pady=4,
            relief="flat",
            cursor="hand2",
            command=self.start,
        )
        self.btn_start.pack(side="left")

        # 停止检测按钮
        self.btn_stop = tk.Button(
            search_frame,
            text="⏹ 停止检测",
            font=("微软雅黑", 10, "bold"),
            bg=danger,
            fg="white",
            padx=18,
            pady=4,
            relief="flat",
            cursor="hand2",
            state="disabled",
            command=self.stop,
        )
        self.btn_stop.pack(side="left", padx=(8, 0))

        # 其他操作按钮
        tk.Button(
            search_frame,
            text="📥 打开Excel",
            font=("微软雅黑", 10, "bold"),
            bg=success,
            fg="white",
            padx=18,
            pady=4,
            relief="flat",
            command=self.open_excel,
        ).pack(side="left", padx=(8, 0))

        tk.Button(
            search_frame,
            text="🖼 截图目录",
            font=("微软雅黑", 9),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=12,
            pady=4,
            relief="flat",
            command=lambda: os.startfile(str(SCREENSHOTS)),
        ).pack(side="left", padx=(8, 0))

        # AI批量验证按钮（检测完成后可用）
        self.btn_ai = tk.Button(
            search_frame,
            text="🤖 AI深度验证",
            font=("微软雅黑", 9, "bold"),
            bg="#8e44ad",
            fg="white",
            padx=14,
            pady=4,
            relief="flat",
            cursor="hand2",
            state="disabled",
            command=self.ai_verify,
        )
        self.btn_ai.pack(side="left", padx=(8, 0))

        tk.Button(
            search_frame,
            text="📂 项目目录",
            font=("微软雅黑", 9),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=12,
            pady=4,
            relief="flat",
            command=lambda: os.startfile(str(BASE)),
        ).pack(side="left", padx=(8, 0))

        # 进度标签
        self.prog = tk.Label(search_frame, text="", font=("微软雅黑", 9), bg="white", fg="#7f8c8d", anchor="w")
        self.prog.pack(side="right", fill="x", expand=True, padx=(10, 0))

        # ========== 登录提示栏 ==========
        login_tip_frame = tk.Frame(main, bg="#fff8e1", highlightbackground="#f0dca0", highlightthickness=1, padx=10, pady=4)
        login_tip_frame.pack(fill="x", pady=(0, 6))
        tk.Label(
            login_tip_frame,
            text="💡 登录提示：如未提前登录淘宝，建议在浏览器弹出后使用「扫码登录」方式，短信登录可能被运营商拦截导致失败。登录等待超时1分钟，请尽快完成登录。",

            font=("微软雅黑", 9),
            bg="#fff8e1",
            fg="#8d6e00",
            anchor="w",
            justify="left",
        ).pack(fill="x")


        # ========== 统计卡片 ==========
        stats = tk.Frame(main, bg="#f5f7fa")
        stats.pack(fill="x", pady=(0, 6))
        self.st = {}
        for label, color in [("总商品", "#2c3e50"), ("疑似侵权", "#e74c3c"), ("书籍排除", "#27ae60")]:
            frame = tk.Frame(stats, bg="white", highlightbackground="#dde", highlightthickness=1, padx=14, pady=6)
            value = tk.Label(frame, text="-", font=("微软雅黑", 18, "bold"), bg="white", fg=color)
            value.pack()
            tk.Label(frame, text=label, font=("微软雅黑", 8), bg="white", fg="#7f8c8d").pack()
            frame.pack(side="left", fill="x", expand=True, padx=(0, 4))
            self.st[label] = value

        # ========== 结果表格 ==========
        table_frame = tk.Frame(main, bg="white", highlightbackground="#dde", highlightthickness=1)
        table_frame.pack(fill="both", expand=True)

        cols = ("#", "商品名称", "价格", "侵权", "时间")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=16)
        for col in cols:
            self.tree.heading(col, text=col)
        self.tree.column("#", width=35, anchor="center")
        self.tree.column("商品名称", width=580)
        self.tree.column("价格", width=80, anchor="center")
        self.tree.column("侵权", width=90, anchor="center")
        self.tree.column("时间", width=130, anchor="center")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.tag_configure("inf", background="#fff0f0")
        self.tree.bind("<Double-1>", self._open)

        # ========== 底部计数 ==========
        self.cnt = tk.Label(main, text="", font=("微软雅黑", 9), bg="#f5f7fa", fg="#7f8c8d")
        self.cnt.pack(anchor="w", pady=(2, 0))

    def _set(self, text, bg=None):
        self.slb.config(text=text)
        if bg:
            self.slb.config(bg=bg)
        self.root.update_idletasks()

    def start(self):
        pages_str = self.pages_var.get().strip()
        if not pages_str:
            messagebox.showwarning("提示", "请输入搜索页数")
            return
        try:
            max_pages = int(pages_str)
            if max_pages < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning("提示", "页数必须为正整数")
            return

        self._running = True
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.pages_entry.config(state="disabled")

        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)
        for value in self.st.values():
            value.config(text="-")
        self.cnt.config(text="")
        self.prog.config(text="")
        self._set("检测中...", "#f39c12")

        threading.Thread(target=self._run, args=(max_pages,), daemon=True).start()

    def stop(self):
        """停止检测"""
        if self._detector:
            self._detector.stop()
        self._running = False
        self.btn_stop.config(state="disabled")
        self.prog.config(text="正在停止...")

    def _run(self, max_pages):
        try:
            def cb(msg):
                if not self._running:
                    return
                self.root.after(0, lambda m=msg: self.prog.config(text=m))

            self._detector = Detector(progress_cb=cb, max_pages=max_pages)
            result = self._detector.run()
            if len(result) == 5:
                results, inf_cnt, book_cnt, total_scanned, excel_path = result
            elif len(result) == 4:
                results, inf_cnt, book_cnt, excel_path = result
                total_scanned = len(results)
            else:
                results, inf_cnt, book_cnt = result[:3]
                total_scanned = len(results)
                excel_path = ""

            self.root.after(0, lambda: self._done(results, inf_cnt, book_cnt, total_scanned, excel_path))
        except StopIteration:
            self.root.after(0, lambda: self._stopped())
        except Exception as exc:
            detail = traceback.format_exc()
            self.root.after(0, lambda: self._failed(exc, detail))

    def _done(self, results, inf_cnt, book_cnt, total_scanned, excel_path):
        self._running = False
        self.results = results
        self._last_excel = excel_path or ""
        self.st["总商品"].config(text=str(total_scanned))
        self.st["疑似侵权"].config(text=str(inf_cnt))
        self.st["书籍排除"].config(text=str(book_cnt))

        for r in results:
            record_time = r.get("记录时间", "")
            display_time = record_time.split(" ")[1] if " " in record_time else record_time
            self.tree.insert(
                "",
                "end",
                values=(
                    r.get("序号", ""),
                    r.get("商品名称", "")[:80],
                    r.get("价格", ""),
                    r.get("是否侵权", ""),
                    display_time,
                ),
                tags=("inf",),
            )

        self.cnt.config(text=f"共 {len(results)} 个 · 侵权 {inf_cnt} · 排除书籍 {book_cnt}")
        self.btn_start.config(state="normal", text="🚀 重新检测")
        self.btn_stop.config(state="disabled")
        self.pages_entry.config(state="normal")
        # 检测完成后启用AI验证按钮
        if AI_BATCH_AVAILABLE and results:
            self.btn_ai.config(state="normal")
        self._set("完成 ✅", "#27ae60")
        self.prog.config(text="检测完成！")


    def _stopped(self):
        self._running = False
        self.btn_start.config(state="normal", text="🚀 重新检测")
        self.btn_stop.config(state="disabled")
        self.pages_entry.config(state="normal")
        self._set("已停止", "#e74c3c")
        self.prog.config(text="🛑 检测已停止")

    def _failed(self, exc, detail):
        self._running = False
        self.btn_start.config(state="normal", text="🚀 重新检测")
        self.btn_stop.config(state="disabled")
        self.pages_entry.config(state="normal")
        self._set("出错", "#e74c3c")
        self.prog.config(text=str(exc))
        messagebox.showerror("检测失败", detail)

    def _open(self, evt):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(self.tree.item(sel[0], "values")[0]) - 1
        if 0 <= idx < len(self.results):
            url = self.results[idx].get("商品URL", "")
            if url:
                webbrowser.open(url)

    def ai_verify(self):
        """AI批量深度验证"""
        if not self.results:
            messagebox.showwarning("提示", "没有检测结果可供验证")
            return

        if not AI_BATCH_AVAILABLE:
            messagebox.showerror("错误", "AI验证模块未加载，请检查 backend/ai_verifier.py 是否存在")
            return

        # 禁用按钮，防止重复点击
        self.btn_ai.config(state="disabled", text="⏳ AI验证中...")
        self._set("AI验证中...", "#8e44ad")
        self.prog.config(text="正在调用DeepSeek API进行AI深度验证...")

        threading.Thread(target=self._run_ai_verify, daemon=True).start()

    def _run_ai_verify(self):
        """在后台线程中执行AI批量验证"""
        try:
            def cb(msg):
                self.root.after(0, lambda m=msg: self.prog.config(text=m))

            # 调用批量验证
            verified_results = batch_verify(self.results, progress_cb=cb)

            # 在主线程更新UI
            self.root.after(0, lambda: self._ai_done(verified_results))
        except Exception as exc:
            detail = traceback.format_exc()
            self.root.after(0, lambda: self._ai_failed(exc, detail))

    def _ai_done(self, verified_results):
        """AI验证完成后的UI更新"""
        self.results = verified_results
        self.btn_ai.config(state="normal", text="🤖 AI深度验证")

        # 统计AI验证结果
        ai_infringing = sum(1 for r in verified_results if r.get("ai_result", {}).get("is_infringement") is True)
        ai_clean = sum(1 for r in verified_results if r.get("ai_result", {}).get("is_infringement") is False)
        ai_unknown = sum(1 for r in verified_results if r.get("ai_result", {}).get("is_infringement") is None)

        # 更新表格 - 添加AI验证结果列
        # 由于ttk.Treeview动态加列比较麻烦，我们在商品名称后追加AI标记
        for i, item in enumerate(self.tree.get_children()):
            r = verified_results[i] if i < len(verified_results) else None
            if r and r.get("ai_result"):
                ai = r["ai_result"]
                ai_inf = ai.get("is_infringement")
                conf_score = ai.get("confidence_score", 0)
                ai_sug = ai.get("suggestion", "")

                # 在商品名称后追加AI标记
                current_vals = list(self.tree.item(item, "values"))
                name = current_vals[1] if len(current_vals) > 1 else ""
                if ai_inf is True:
                    marker = f" [🤖侵权-{conf_score}分]"
                elif ai_inf is False:
                    marker = f" [🤖非侵权-{conf_score}分]"
                else:
                    marker = f" [🤖未知-{conf_score}分]"

                if len(current_vals) > 1:
                    current_vals[1] = (name + marker)[:80]
                    self.tree.item(item, values=tuple(current_vals))


        # 更新统计信息
        ai_summary = f"AI验证: {ai_infringing}侵权 / {ai_clean}非侵权 / {ai_unknown}未知"
        self.cnt.config(text=f"共 {len(verified_results)} 个 · {ai_summary}")
        self._set("AI验证完成 ✅", "#8e44ad")
        self.prog.config(text=f"🤖 AI深度验证完成！侵权:{ai_infringing} 非侵权:{ai_clean} 未知:{ai_unknown}")

        # 提示用户重新生成Excel以包含AI结果
        if messagebox.askyesno("AI验证完成", f"AI验证完成！\n\n侵权: {ai_infringing}\n非侵权: {ai_clean}\n未知: {ai_unknown}\n\n是否重新生成包含AI验证结果的Excel报告？"):
            self._regenerate_excel_with_ai()

    def _ai_failed(self, exc, detail):
        """AI验证失败处理"""
        self.btn_ai.config(state="normal", text="🤖 AI深度验证")
        self._set("AI验证失败", "#e74c3c")
        self.prog.config(text=f"AI验证失败: {exc}")
        messagebox.showerror("AI验证失败", f"错误: {exc}\n\n详情:\n{detail}")

    def _regenerate_excel_with_ai(self):
        """重新生成包含AI验证结果的Excel报告（带时间戳，与原始报告对应）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XlImage
            import openpyxl.utils
            import datetime

            # AI验证报告单独存放在 ai_reports 目录（带时间戳，与原始报告对应）
            AI_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = AI_REPORTS_DIR / f"侵权检测报告_AI验证_{timestamp}.xlsx"




            wb = Workbook()
            ws = wb.active
            ws.title = "侵权检测报告(AI验证)"

            headers = ["序号","商品名称","商品截图","商品URL","记录时间","价格","是否侵权","AI侵权校验","AI置信度","AI建议"]
            hf = Font(bold=True, size=11, color="FFFFFF")
            hb = PatternFill("solid", fgColor="4472C4")
            ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font=hf; c.fill=hb; c.alignment=ha; c.border=thin

            widths = [8,50,22,55,20,12,12,12,12,18]
            for ci,w in enumerate(widths,1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

            # 在表头下方插入AI判断标准说明
            note_row = 2
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
            note_text = (
                "【AI侵权判断标准】"
                " 评分制(满分100分)：①商标使用(30分)含校名/缩写/CUC ②商品类别(25分)服装/文具/饰品等周边类 "
                "③官方关联暗示(20分)官方/正版/授权/纪念/周边/文创 ④价格异常(15分)低价高分 ⑤卖家非官方(10分)非北京商家高分。"
                " ≥80分确认侵权，60-79分酌情复核，<60分建议复核。核心原则：非书籍教材类商品含校名+周边类别=侵权。"
            )
            c_note = ws.cell(row=note_row, column=1, value=note_text)
            c_note.font = Font(color="555555", size=9, italic=True)
            c_note.alignment = Alignment(vertical="center", wrap_text=True)
            c_note.fill = PatternFill("solid", fgColor="FFF8E1")
            ws.row_dimensions[note_row].height = 50

            black = Font(color="000000")
            red = Font(color="FF0000", bold=True)
            green = Font(color="27ae60", bold=True)
            orange = Font(color="e67e22", bold=True)

            for ri, r in enumerate(self.results, 3):

                c1 = ws.cell(row=ri, column=1, value=r.get("序号", ''))
                c1.border=thin; c1.alignment=Alignment(vertical="center",wrap_text=True); c1.font = black

                c2 = ws.cell(row=ri, column=2, value=r.get("商品名称", ''))
                c2.border=thin; c2.alignment=Alignment(vertical="center",wrap_text=True); c2.font = black

                img_path = r.get("截图路径", "")
                if img_path and os.path.exists(img_path):
                    try:
                        img = XlImage(img_path)
                        img.width = 120
                        img.height = 120 * img.height / img.width if img.width > 0 else 120
                        if img.height > 160: img.height = 160
                        ws.add_image(img, f"C{ri}")
                        ws.row_dimensions[ri].height = max(ws.row_dimensions[ri].height or 0, img.height + 4)
                    except Exception:
                        pass

                c4 = ws.cell(row=ri, column=4, value=r.get("商品URL", ''))
                c4.border=thin; c4.alignment=Alignment(vertical="center",wrap_text=True)
                url = r.get("商品URL", '')
                if url:
                    c4.hyperlink = url
                    c4.font = Font(color="0563C1", underline="single")

                c5 = ws.cell(row=ri, column=5, value=r.get("记录时间", ''))
                c5.border=thin; c5.alignment=Alignment(vertical="center",wrap_text=True); c5.font = black

                c6 = ws.cell(row=ri, column=6, value=r.get("价格", ''))
                c6.border=thin; c6.alignment=Alignment(vertical="center",wrap_text=True); c6.font = black

                c7 = ws.cell(row=ri, column=7, value=r.get("是否侵权", ''))
                c7.border=thin; c7.alignment=Alignment(vertical="center",wrap_text=True); c7.font = red


                # AI验证结果 - 精确到每一分
                ai = r.get("ai_result", {})
                if ai:
                    ai_inf = ai.get("is_infringement")
                    conf_score = ai.get("confidence_score", 0)
                    ai_sug = ai.get("suggestion", "建议复核")

                    # AI侵权校验列：只填"是"或"否"
                    if ai_inf is True:
                        ai_text = "是"
                    elif ai_inf is False:
                        ai_text = "否"
                    else:
                        ai_text = "建议复核"

                    # 根据分数决定建议和颜色
                    if ai_inf is True:
                        if conf_score >= 80:
                            ai_sug = "确认侵权"
                            ai_font = Font(color="FF0000", bold=True)
                        elif conf_score >= 60:
                            ai_sug = "酌情复核"
                            ai_font = Font(color="FF0000", bold=True)
                        else:
                            ai_sug = "建议复核"
                            ai_font = Font(color="FF0000", bold=True)
                    elif ai_inf is False:
                        ai_font = green
                    else:
                        ai_font = orange

                    c8 = ws.cell(row=ri, column=8, value=ai_text)
                    c8.border=thin; c8.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True); c8.font = ai_font

                    c9 = ws.cell(row=ri, column=9, value=f"{conf_score}%")
                    c9.border=thin; c9.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True); c9.font = ai_font

                    c10 = ws.cell(row=ri, column=10, value=ai_sug)
                    c10.border=thin; c10.alignment=Alignment(vertical="center", wrap_text=True); c10.font = ai_font



            wb.save(str(excel_path))
            self._last_excel = str(excel_path)
            self.prog.config(text=f"✅ AI验证报告已生成: {excel_path.name}")

            if messagebox.askyesno("报告已生成", f"AI验证报告已保存:\n{excel_path}\n\n是否立即打开？"):
                os.startfile(str(excel_path))
        except Exception as e:
            messagebox.showerror("生成报告失败", str(e))

    def open_excel(self):
        if self._last_excel and os.path.exists(self._last_excel):
            os.startfile(self._last_excel)
            return

        reports = sorted(REPORTS_DIR.glob("侵权检测报告_*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
        if reports:
            os.startfile(str(reports[0]))
        else:
            messagebox.showwarning("提示", "还没有生成Excel文件")

    def run(self):
        self.root.mainloop()



if __name__ == "__main__":
    App().run()
