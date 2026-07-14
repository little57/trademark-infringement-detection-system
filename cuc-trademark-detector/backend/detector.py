# -*- coding: utf-8 -*-
"""
中国传媒大学 淘宝侵权商品检测 - 完整版
支持：翻页、防爬、书籍过滤、侵权识别、截图、Excel报告
"""
import sys, os, re, json, time, datetime, urllib.parse, pathlib, subprocess, glob, random, threading, logging

# 日志配置
LOG_DIR = pathlib.Path(__file__).resolve().parent.parent / "data" / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "detector.log"
logging.basicConfig(
    filename=str(LOG_FILE),
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8",
)
logger = logging.getLogger("detector")

BASE = pathlib.Path(__file__).resolve().parent.parent
SCREENSHOTS = BASE / "data" / "screenshots"
REPORTS_DIR = BASE / "data" / "reports"

# AI增强验证模块（可选导入，不影响原有功能）
try:
    from backend.ai_verifier import ai_enhanced_judgment, batch_verify
    AI_VERIFIER_AVAILABLE = True
except ImportError:
    AI_VERIFIER_AVAILABLE = False
except Exception:
    AI_VERIFIER_AVAILABLE = False

SCREENSHOTS.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
KEYWORD = "中国传媒大学"
BOOK_KWS = ["书籍","教材","课本","图书","书","教辅","考研","真题","题库","讲义","复习","资料","考试","笔记","辅导","文具","本子","笔",
            "出版社","当当","新华书店","图书专营","书店"]
# 商标侵权关键词：校名、校徽、缩写等
# 注意：只包含明确指向学校的标识性关键词，不含通用商品描述词
TRADEMARK_KWS = ["中国传媒大学","中传","CUC","校徽","校名"]
# 侵权商品类别关键词
INFRINGEMENT_CATEGORY_KWS = ["T恤","卫衣","衣服","服装","外套","夹克","冲锋衣","帽子","鸭舌帽","棒球帽","帆布包","手提袋","包",
                              "笔记本","记事本","珐琅","钥匙扣","挂件","冰箱贴","磁贴",
                              "手机壳","马克杯","水杯","杯子","抱枕","靠垫","坐垫","口罩","书签","明信片",
                              "定制","批发","批量","礼品","礼物","礼盒","礼袋","礼品袋","礼品盒","礼品套装",]
# 低价阈值：低于此价格认为可能是侵权商品
LOW_PRICE_THRESHOLD = 100
STEALTH_JS = """
Object.defineProperty(navigator, 'webdriver', { get: () => false });
Object.defineProperty(navigator, 'plugins', { get: () => [1,2,3,4,5] });
Object.defineProperty(navigator, 'languages', { get: () => ['zh-CN', 'zh', 'en'] });
window.chrome = { runtime: {} };
const origQuery = window.navigator.permissions.query.bind(window.navigator.permissions);
window.navigator.permissions.query = (params) => {
    if (params.name === 'notifications') return Promise.resolve({state: 'denied'});
    return origQuery(params);
};
"""
IS_LOGGED_IN_JS = """
() => {
    // 方式1：检测Cookie中是否有登录标记（最可靠）
    const cookies = document.cookie.split(';').map(c => c.trim().split('=')[0]);
    const loginCookies = ['_tb_token_', 'cookie2', 'l', 'uc1', 'uc3', 'tracknick', 'dnk', 'thw', 'unb'];
    for (const lc of loginCookies) {
        if (cookies.includes(lc)) return true;
    }
    
    // 方式2：检测用户昵称元素（淘宝新版）
    const nickSelectors = [
        '.site-nav-user .site-nav-login-info-nick',
        '.J_MemberNick', '.member-nick', '.tb-member-nick',
        '.site-nav-user .user-nick', '.user-nick',
        '.site-nav-user .nick', '.nick',
        '.site-nav-user a[class*="nick"]',
        '.site-nav-user span[class*="nick"]',
        '#J_UserInfo .nick', '#J_UserInfo a[class*="nick"]',
        '.site-nav-bd .user', '.site-nav-bd a[class*="user"]',
        '.site-nav-user .username', '.username',
        '.site-nav-user .login-info .name',
        '.header-user .name', '.header-user .nick',
        '.top-nav-user .name',
        '.site-nav-user [class*="user"] a',
        '.site-nav-user [class*="member"]',
        '.site-nav-user [class*="login"] span',
    ];
    for (const sel of nickSelectors) {
        const el = document.querySelector(sel);
        if (el && el.textContent && el.textContent.trim() && el.textContent.trim().length > 0 && el.textContent.trim().length < 30) {
            return true;
        }
    }
    
    // 方式3：检测头像元素
    const avatarSelectors = [
        '.site-nav-user .avatar', '.J_Avatar', '.user-avatar',
        '.site-nav-user img[class*="avatar"]', '.site-nav-user img[class*="head"]',
        '.site-nav-user .user-photo', '.user-photo',
        '.header-user .avatar', '.header-user img[class*="avatar"]',
    ];
    for (const sel of avatarSelectors) {
        const el = document.querySelector(sel);
        if (el && el.offsetParent !== null) return true;
    }
    
    // 方式4：检测登录按钮（如果登录按钮可见，说明未登录）
    const loginBtnSelectors = [
        '.site-nav-login-info a[href*="login"]', '.J_Login', '.btn-login',
        '.site-nav-user a[href*="login"]', '.site-nav a[href*="login"]',
        '.top-nav a[href*="login"]', '.header a[href*="login"]',
        'a[href*="login.taobao.com"]',
    ];
    for (const sel of loginBtnSelectors) {
        const el = document.querySelector(sel);
        if (el && el.offsetParent !== null) return false;
    }
    
    // 方式5：检测页面文字
    const bodyText = document.body && document.body.innerText ? document.body.innerText : '';
    if (bodyText.includes('请登录') || bodyText.includes('登录淘宝')) return false;
    if (bodyText.includes('我的淘宝') || bodyText.includes('我的订单') || bodyText.includes('已买到的宝贝')) return true;
    
    return null;
}
"""


CAPTCHA_CHECK_JS = """
() => {
    const captchaModal = document.querySelector('.nc-container, .nc_modal, #nc_1_nc-container, .sm-pop, .sms-popup');
    const captchaSlider = document.querySelector('.nc_iconfont, .btn_slide, .nc-slider, .slider-icon');
    const hasCaptchaIframe = document.querySelector('iframe[src*="captcha"], iframe[src*="nc"], iframe[src*="verify"]');
    const textMatch = document.body.innerText && (
        document.body.innerText.includes('请拖动下方滑块完成验证') ||
        document.body.innerText.includes('请按住滑块，拖动到最右边') ||
        document.body.innerText.includes('滑动验证') ||
        document.body.innerText.includes('安全验证')
    );
    return !!(captchaModal || captchaSlider || hasCaptchaIframe || textMatch);
}
"""
DISMISS_CAPTCHA_JS = """
() => {
    const closeBtns = document.querySelectorAll('.nc_close, .sm-close, .sms-close, .close-btn, .nc-icon-close, [class*="close"][class*="nc"], .smt-close, .btn-close');
    for (const btn of closeBtns) { if (btn.offsetParent !== null) { btn.click(); return true; } }
    const masks = document.querySelectorAll('.nc_mask, .sm-mask, .sms-mask');
    for (const mask of masks) { if (mask.offsetParent !== null) { mask.click(); return true; } }
    document.dispatchEvent(new KeyboardEvent('keydown', {key: 'Escape', code: 'Escape', keyCode: 27, which: 27}));
    return false;
}
"""
LOGIN_CHECK_JS = """
() => {
    const loginModal = document.querySelector('.login-pop, .login-dialog, #login, .J_LoginDialog, .tb-login');
    const loginIframe = document.querySelector('iframe[src*="login.taobao.com"], iframe[src*="login.alipay"]');
    const loginBox = document.getElementById('J_LoginBox');
    const isLoginUrl = window.location.href.includes('login.taobao.com') || window.location.href.includes('login.alipay.com');
    return !!(loginModal || loginIframe || loginBox || isLoginUrl);
}
"""
DISMISS_LOGIN_JS = """
() => {
    const closeBtns = document.querySelectorAll('.login-close, .close-btn, .J_Close, .pop-close, .dialog-close, [class*="close"][class*="login"], .icon-close, .smt-close, .tb-btn-close');
    for (const btn of closeBtns) { if (btn.offsetParent !== null) { btn.click(); return true; } }
    document.dispatchEvent(new KeyboardEvent('keydown', {key: 'Escape', code: 'Escape', keyCode: 27, which: 27}));
    return false;
}
"""
# 翻页JS：点击页码按钮翻页
CLICK_PAGE_JS = """
(targetPage) => {
    // 方式1：点击页码链接
    const pageLinks = document.querySelectorAll('.pagination a, .page-item a, a[class*="page"], a[class*="Page"], .J_Page a, .J_Ajax a');
    for (const link of pageLinks) {
        const text = (link.textContent || '').trim();
        if (text === String(targetPage)) {
            link.click();
            return true;
        }
    }
    // 方式2：点击页码按钮
    const pageBtns = document.querySelectorAll('.pagination button, .page-item button, button[class*="page"], button[class*="Page"]');
    for (const btn of pageBtns) {
        const text = (btn.textContent || '').trim();
        if (text === String(targetPage)) {
            btn.click();
            return true;
        }
    }
    // 方式3：从所有页码元素中找
    const allPageEls = document.querySelectorAll('[class*="page"] a, [class*="Page"] a, [class*="pagination"] span, [class*="Pagination"] span');
    for (const el of allPageEls) {
        const text = (el.textContent || '').trim();
        if (text === String(targetPage) && el.offsetParent !== null) {
            el.click();
            return true;
        }
    }
    return false;
}
"""
GET_PAGE_INFO_JS = """
() => {
    // 获取当前页码和总页数
    const pageLinks = document.querySelectorAll('.pagination a, .page-item a, a[class*="page"], a[class*="Page"], .J_Page a, .J_Ajax a');
    const pages = [];
    let currentPage = 1;
    for (const link of pageLinks) {
        const text = (link.textContent || '').trim();
        const num = parseInt(text);
        if (!isNaN(num) && num > 0 && num < 100) {
            pages.push(num);
        }
        if (link.classList.contains('active') || link.classList.contains('current') || link.parentElement.classList.contains('active')) {
            currentPage = num;
        }
    }
    // 也检查span
    const pageSpans = document.querySelectorAll('.pagination span, .page-item span, .active, .current');
    for (const span of pageSpans) {
        const text = (span.textContent || '').trim();
        const num = parseInt(text);
        if (!isNaN(num) && num > 0 && num < 100) {
            if (span.classList.contains('active') || span.classList.contains('current') || span.parentElement.classList.contains('active')) {
                currentPage = num;
            }
        }
    }
    const maxPage = pages.length > 0 ? Math.max(...pages) : 1;
    return { currentPage, maxPage, availablePages: [...new Set(pages)].sort((a,b) => a-b) };
}
"""
EXTRACT_ITEMS_JS = """
() => {
    // 辅助函数：从文本中提取价格，只保留2位小数
    function extractPrice(text) {
        if (!text) return '';
        // 匹配价格模式：数字开头，可能带小数点
        const matches = text.match(/(\\d+\\.?\\d*)/);
        if (matches) {
            let p = matches[1];
            // 强制只保留2位小数
            const dotIdx = p.indexOf('.');
            if (dotIdx > 0) {
                p = p.substring(0, dotIdx + Math.min(p.length - dotIdx, 3));
            }
            return p;
        }
        return '';
    }


    const list = [];
    // 方式1：从g_page_config全局变量
    try { if (typeof g_page_config !== 'undefined' && g_page_config.mods?.itemlist?.data?.auctions) {
        g_page_config.mods.itemlist.data.auctions.forEach(a => { list.push({
            title: (a.raw_title||a.title||'').replace(/<[^>]+>/g,''),
            url: 'https:' + (a.detail_url||''),
            price: extractPrice(a.view_price||''),
            seller: a.nick||'',
            location: a.item_loc||'',
        }); });
    } } catch(e) {}
    // 方式2：从script标签
    if (list.length === 0) { for (const s of document.querySelectorAll('script')) {
        const m = (s.textContent||'').match(/g_page_config\\s*=\\s*({.*?});/);
        if (m) try { const cfg = JSON.parse(m[1]); (cfg.mods?.itemlist?.data?.auctions||[]).forEach(a => list.push({
            title: (a.raw_title||a.title||'').replace(/<[^>]+>/g,''),
            url: 'https:' + (a.detail_url||''),
            price: extractPrice(a.view_price||''),
            seller: a.nick||'',
            location: a.item_loc||'',
        })); } catch(e) {}
    } }
    // 方式3：从DOM元素
    if (list.length === 0) {
        document.querySelectorAll('[data-spm*="item"], .search-item, .item-card, [class*="Card"]').forEach(el => { try {
            const link = el.querySelector('a[href*="item.taobao.com"], a[href*="detail.tmall.com"]');
            const titleEl = el.querySelector('[class*="Title"], [class*="title"], [class*="Name"], [class*="name"]');
            const priceEl = el.querySelector('[class*="Price"], [class*="price"]');
            const shopEl = el.querySelector('[class*="Shop"], [class*="shop"], [class*="Seller"], [class*="seller"]');
            const locEl = el.querySelector('[class*="location"], [class*="Location"], [class*="loc"], [class*="address"]');
            const title = titleEl ? (titleEl.textContent || titleEl.innerText || '').trim() : '';
            const price = priceEl ? extractPrice(priceEl.textContent || '') : '';
            const seller = shopEl ? (shopEl.textContent || '').trim() : '';
            const location = locEl ? (locEl.textContent || '').trim() : '';
            let url = link ? (link.href || '') : '';
            if (url && url.startsWith('//')) url = 'https:' + url;
            if (title && url && title.length > 2) list.push({ title, url, price, seller, location });
        } catch(e) {} });
    }
    // 方式4：从所有商品链接
    if (list.length === 0) { const seen = new Set();
        document.querySelectorAll('a[href*="item.taobao.com"], a[href*="detail.tmall.com"]').forEach(link => { try {
            let url = link.href || ''; if (url && url.startsWith('//')) url = 'https:' + url;
            if (seen.has(url)) return; seen.add(url);
            let parent = link.closest('[class*="item"], [class*="card"], [class*="Card"], li, .search-item') || link.parentElement;
            let title = link.textContent || link.title || '';
            if (!title || title.length < 3) { title = parent ? (parent.textContent || '').trim() : ''; title = title.substring(0, 120); }
            let price = '';
            if (parent) { const pEl = parent.querySelector('[class*="price"], [class*="Price"]'); if (pEl) price = extractPrice(pEl.textContent || ''); }
            let location = '';
            if (parent) { const lEl = parent.querySelector('[class*="location"], [class*="Location"]'); if (lEl) location = (lEl.textContent || '').trim(); }
            if (title && title.length > 2) list.push({ title: title.replace(/<[^>]+>/g,'').trim(), url, price, seller: '', location });
        } catch(e) {} });
    }
    const seen = new Set(); const uniqueList = [];
    for (const item of list) { if (!seen.has(item.url)) { seen.add(item.url); uniqueList.push(item); } }
    return uniqueList;
}
"""


def random_sleep(min_sec, max_sec):
    """随机休眠一段时间，用于模拟人类行为，避免被反爬"""
    delay = random.uniform(min_sec, max_sec)
    time.sleep(delay)

def is_book(title):
    t = title.lower()
    return any(k in t for k in BOOK_KWS)

def is_in_beijing(location):
    """判断商品所在地是否在北京"""
    if not location:
        return False
    return '北京' in location

def has_trademark_keyword(title):
    """标题是否包含校名校徽等商标关键词"""
    t = title.lower()
    return any(k.lower() in t for k in TRADEMARK_KWS)

def has_infringement_category(title):
    """标题是否属于侵权商品类别（服装、文具、饰品等）"""
    t = title.lower()
    return any(k.lower() in t for k in INFRINGEMENT_CATEGORY_KWS)

def is_low_price(price_str):
    """判断价格是否低于低价阈值"""
    try:
        price = float(price_str)
        return price < LOW_PRICE_THRESHOLD
    except (ValueError, TypeError):
        return False

def is_suspected_infringement(title, price, location, use_ai=False, progress_cb=None):
    """
    商标侵权判定逻辑（五步核查）+ 可选AI增强
    
    参数:
        title: 商品标题
        price: 价格
        location: 所在地
        use_ai: 是否启用AI增强判断（混合模式）
        progress_cb: 进度回调（用于AI调用时显示进度）
    
    返回:
        (是否侵权, 判定理由, ai_used)
        - ai_used: 是否使用了AI辅助判断
    """
    reasons = []
    
    # 步骤1：是否商业使用（标题含校名校徽+上架售卖）
    has_tm = has_trademark_keyword(title)
    if not has_tm:
        return False, "标题不含校名校徽关键词", False
    reasons.append("标题含校名校徽关键词")
    
    # 步骤2：比对标识是否相同/近似
    # 标题含"中国传媒大学"、"CUC"、"中传"、"校徽"等，视为相同/近似
    if "中国传媒大学" in title or "中传" in title:
        reasons.append("含完整校名/简称")
    if "CUC" in title.upper():
        reasons.append("含CUC缩写")
    if "校徽" in title:
        reasons.append("含校徽标识")
    
    # 步骤3：商品类别是否匹配（服装、文具、饰品等）
    has_cat = has_infringement_category(title)
    if not has_cat:
        return False, "商品类别不在侵权保护范围内", False
    reasons.append("商品类别匹配")
    
    # 步骤4：核查商家授权（无授权推定侵权）
    # 非北京商家 + 低价 = 无授权可能性大
    is_bj = is_in_beijing(location)
    if is_bj:
        reasons.append("北京商家（需人工核查授权）")
    else:
        reasons.append("非北京商家（推定无授权）")
    
    # 步骤5：是否造成混淆
    # 标题含"纪念"、"周边"、"文创"、"官方"等词，易造成混淆
    confusion_kws = ["纪念","周边","文创","官方","纪念品","纪念款"]
    if any(k in title for k in confusion_kws):
        reasons.append("易造成官方混淆")
    
    # 综合判定
    # 核心条件：有商标关键词 + 有商品类别匹配 = 侵权
    # 辅助条件：非北京 + 低价 = 高置信度侵权
    confidence = "高" if (not is_bj and is_low_price(price)) else ("中" if not is_bj else "低")
    reasons.append(f"置信度:{confidence}")
    
    rule_result = True
    rule_reason = "; ".join(reasons)
    
    # ========== AI增强判断（混合模式） ==========
    ai_used = False
    if use_ai and AI_VERIFIER_AVAILABLE:
        try:
            final_result, final_reason, ai_used = ai_enhanced_judgment(
                title=title,
                price=price,
                location=location,
                rule_result=rule_result,
                rule_reason=rule_reason,
                progress_cb=progress_cb,
            )
            return final_result, final_reason, ai_used
        except Exception:
            # AI调用失败，回退到规则结果
            pass
    
    return rule_result, rule_reason, ai_used


def detect_browser():
    if os.name == 'nt':
        edge_paths = ["C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe", "C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe"]
        chrome_paths = ["C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe", "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"]
        for p in edge_paths:
            if os.path.exists(p): return "edge", p
        for p in chrome_paths:
            if os.path.exists(p): return "chrome", p
    return "chromium", None

COOKIE_FILE = BASE / "data" / "taobao_cookies.json"

def load_cookies(ctx):
    if not COOKIE_FILE.exists(): return False
    try:
        with open(COOKIE_FILE, "r", encoding="utf-8") as f:
            cookies = json.load(f)
        for c in cookies:
            c.pop('sameSite', None); c.pop('priority', None); c.pop('sameParty', None); c.pop('sourceScheme', None); c.pop('sourcePort', None)
        ctx.add_cookies(cookies)
        return True
    except: return False

def save_cookies(ctx):
    try:
        cookies = ctx.cookies()
        COOKIE_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(COOKIE_FILE, "w", encoding="utf-8") as f:
            json.dump(cookies, f, ensure_ascii=False, indent=2)
        return True
    except: return False

def _handle_login(page, max_retries=3):
    for i in range(max_retries):
        try:
            is_login = page.evaluate(LOGIN_CHECK_JS)
            if not is_login: return False
            self_progress = getattr(page, '_progress', None) or (lambda m: None)
            self_progress("检测到登录弹窗，尝试关闭...")
            dismissed = page.evaluate(DISMISS_LOGIN_JS)
            page.wait_for_timeout(1000)
            if dismissed: self_progress("已关闭登录弹窗"); return True
        except: pass
        page.wait_for_timeout(500)
    return True

def _check_login_by_cookies(ctx):
    """通过Cookie检测是否已登录淘宝（最可靠方式）"""
    try:
        cookies = ctx.cookies()
        cookie_names = {c['name'] for c in cookies}
        login_cookies = {'_tb_token_', 'cookie2', 'l', 'uc1', 'uc3', 'tracknick', 'dnk', 'thw', 'unb'}
        matched = cookie_names & login_cookies
        if matched:
            logger.info(f"Cookie检测到登录标记: {matched}")
            return True
        return False
    except Exception as e:
        logger.warning(f"Cookie检测异常: {e}")
        return False

def _wait_for_login(page, ctx, timeout_minutes=10, stop_event=None):
    self_progress = getattr(page, '_progress', None) or (lambda m: None)
    self_progress("⚠️ 请在浏览器中手动登录淘宝（推荐扫码登录，短信登录可能被拦截）")
    self_progress(f"⏳ 等待登录中（最长{timeout_minutes}分钟）...")
    logger.info("开始等待登录...")
    start = time.time()
    while time.time() - start < timeout_minutes * 60:
        # 检查是否被用户终止
        if stop_event and stop_event.is_set():
            self_progress("🛑 用户终止等待登录")
            logger.info("用户终止等待登录")
            return False
        try:
            # 检查页面URL和状态
            current_url = page.url
            logger.debug(f"等待登录中 - 当前URL: {current_url}")
            
            # 检查页面是否还活着
            closed = page.evaluate("document.body === null")
            if closed:
                self_progress("⚠️ 页面已关闭，尝试重新导航到淘宝...")
                logger.warning("页面body为null，尝试重新导航")
                try:
                    page.goto("https://www.taobao.com", wait_until="domcontentloaded", timeout=15000)
                    page.wait_for_timeout(2000)
                    logger.info("重新导航到淘宝首页成功")
                except Exception as nav_e:
                    logger.error(f"重新导航失败: {nav_e}")
                    self_progress("❌ 页面无法恢复，登录失败")
                    return False
            
            # 方式1：Cookie检测（最可靠）
            if _check_login_by_cookies(ctx):
                self_progress("✅ 登录成功！等待页面稳定...")
                logger.info("登录成功（Cookie检测）！等待5秒稳定...")
                page.wait_for_timeout(5000)
                return True
            
            # 方式2：JS DOM检测（备选）
            logged_in = page.evaluate(IS_LOGGED_IN_JS)
            has_login_modal = page.evaluate(LOGIN_CHECK_JS)
            logger.debug(f"登录状态检查 - logged_in={logged_in}, has_login_modal={has_login_modal}")
            if logged_in is True and not has_login_modal:
                self_progress("✅ 登录成功！等待页面稳定...")
                logger.info("登录成功（DOM检测）！等待5秒稳定...")
                page.wait_for_timeout(5000)
                return True

        except Exception as e:
            err_str = str(e)
            logger.warning(f"等待登录异常: {err_str[:200]}")
            # 页面被关闭或导航到无效页面
            if "closed" in err_str.lower() or "detached" in err_str.lower() or "target" in err_str.lower():
                self_progress("⚠️ 浏览器页面状态异常，尝试重新导航到淘宝...")
                logger.warning("页面状态异常，尝试重新导航")
                try:
                    page.goto("https://www.taobao.com", wait_until="domcontentloaded", timeout=15000)
                    page.wait_for_timeout(2000)
                    logger.info("重新导航到淘宝首页成功")
                except Exception as nav_e:
                    logger.error(f"重新导航失败: {nav_e}")
                    self_progress("❌ 页面无法恢复，登录失败")
                    return False
        page.wait_for_timeout(3000)
    logger.warning("登录超时")
    self_progress("⏰ 登录超时"); return False




def check_login_status(page):
    try:
        _handle_login(page); page.wait_for_timeout(500)
        logged_in = page.evaluate(IS_LOGGED_IN_JS)
        if logged_in is True: return True
        if logged_in is False: return False
        has_login_modal = page.evaluate(LOGIN_CHECK_JS)
        if not has_login_modal and ('taobao.com' in page.url or 'tmall.com' in page.url): return True
        return False
    except: return False

def check_captcha(page):
    try: return page.evaluate(CAPTCHA_CHECK_JS)
    except: return False

def handle_captcha(page, max_retries=3):
    self_progress = getattr(page, '_progress', None) or (lambda m: None)
    for i in range(max_retries):
        try:
            if not check_captcha(page): return True
            self_progress(f"🔐 检测到滑块验证，尝试自动关闭 ({i+1}/{max_retries})...")
            dismissed = page.evaluate(DISMISS_CAPTCHA_JS)
            page.wait_for_timeout(1500)
            if dismissed and not check_captcha(page): self_progress("✅ 滑块验证已关闭"); return True
        except: pass
        page.wait_for_timeout(1000)
    return False

def wait_for_captcha_solve(page, timeout_minutes=5):
    self_progress = getattr(page, '_progress', None) or (lambda m: None)
    self_progress("⚠️ 检测到淘宝滑块验证！")
    self_progress("👉 请在浏览器中手动拖动滑块完成验证")
    self_progress(f"⏳ 等待验证中（最长{timeout_minutes}分钟）...")
    start = time.time()
    while time.time() - start < timeout_minutes * 60:
        try:
            has_captcha = check_captcha(page)
            if not has_captcha:
                page.wait_for_timeout(1000)
                if not check_captcha(page):
                    self_progress("✅ 滑块验证通过！")
                    page.reload(wait_until="domcontentloaded")
                    page.wait_for_timeout(1500)
                    return True
        except:
            pass
        page.wait_for_timeout(2000)
    self_progress("⏰ 滑块验证超时"); return False

class Detector:
    def __init__(self, progress_cb=None, max_pages=10):
        self.progress = progress_cb or (lambda msg: print(msg))
        self._stop_event = threading.Event()
        self._browser = None
        self._ctx = None
        self.max_pages = max_pages
    
    def stop(self):
        self._stop_event.set()
        self.progress("🛑 收到终止指令，正在停止检测...")
    
    def _check_stop(self):
        if self._stop_event.is_set():
            raise StopIteration("检测任务已被用户终止")
    
    def _navigate_and_handle_login(self, page, url, **kwargs):
        try:
            page.goto(url, **kwargs)
            page.wait_for_timeout(2000)
            _handle_login(page)
            if check_captcha(page):
                if not handle_captcha(page):
                    wait_for_captcha_solve(page)
        except Exception as e:
            self.progress(f"导航失败: {e}")
    
    def run(self) -> tuple[list, int, int, int, str]:
        if os.name == 'nt':
            browser_name, browser_path = detect_browser()
            self.progress(f"检测到浏览器: {browser_name}")
        else:
            browser_name = "chromium"
        
        self.progress("启动浏览器...")
        from playwright.sync_api import sync_playwright
        
        with sync_playwright() as p:
            if browser_name == "edge":
                browser = p.chromium.launch(headless=False, channel="msedge", args=["--no-sandbox", "--disable-blink-features=AutomationControlled"])
            elif browser_name == "chrome":
                browser = p.chromium.launch(headless=False, channel="chrome", args=["--no-sandbox", "--disable-blink-features=AutomationControlled"])
            else:
                browser = p.chromium.launch(headless=False, args=["--no-sandbox", "--disable-blink-features=AutomationControlled"])
            
            self._browser = browser
            self._ctx = ctx = browser.new_context(
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                locale="zh-CN", timezone_id="Asia/Shanghai",
            )
            page = ctx.new_page()
            page.add_init_script(STEALTH_JS)
            page._progress = self.progress
            
            # ========== 第一步：登录 ==========
            self.progress("=" * 60)
            self.progress("第一步：登录验证")
            self.progress("=" * 60)
            
            if COOKIE_FILE.exists():
                self.progress("📂 加载本地登录凭证...")
                loaded = load_cookies(ctx)
                self.progress("✅ 登录凭证已加载" if loaded else "⚠️ 加载失败")
            
            self.progress("🌐 访问淘宝首页...")
            self._navigate_and_handle_login(page, "https://www.taobao.com", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(2000)
            
            if not check_login_status(page):
                self.progress("⚠️ 请手动登录淘宝...")
                try:
                    login_btn = page.query_selector('.site-nav-login-info a, .J_Login, .btn-login')
                    if login_btn: login_btn.click(); page.wait_for_timeout(1000)
                except: pass
                if not _wait_for_login(page, ctx, timeout_minutes=1, stop_event=self._stop_event):
                    self.progress("❌ 登录失败"); return [], 0, 0, 0, ""



                save_cookies(ctx)
            
            # ========== 第二步：搜索 + 翻页 + 侵权识别 ==========
            self.progress("=" * 60)
            self.progress("第二步：搜索商品，商标侵权智能判定")
            self.progress("=" * 60)
            
            self.progress(f"🔍 搜索关键词: \"{KEYWORD}\"")
            self.progress(f"📄 最多翻页数: {self.max_pages} 页")
            
            results = []
            page_num = 0
            total_scanned = 0
            total_books = 0
            total_infringing = 0
            
            first_search_url = "https://s.taobao.com/search?q=" + urllib.parse.quote(KEYWORD)
            self.progress("🌐 访问搜索结果页...")
            self._navigate_and_handle_login(page, first_search_url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3000)
            
            # 记录已访问过的页码，避免重复
            visited_pages = set()
            
            while page_num < self.max_pages:
                # 检查是否被用户终止
                if self._stop_event and self._stop_event.is_set():
                    self.progress("🛑 用户终止检测")
                    logger.info("用户终止检测（翻页循环）")
                    break

                current_page_display = page_num + 1
                visited_pages.add(current_page_display)
                self.progress(f"📄 第 {current_page_display} 页...")

                
                # 翻页（优先点击页码按钮，随机跳页）
                if page_num > 0:
                    random_sleep(2, 5)
                    
                    # 先获取当前页的页码信息，看看有哪些页码可用
                    page_info = page.evaluate(GET_PAGE_INFO_JS)
                    available = page_info.get('availablePages', [])
                    
                    # 过滤掉已访问过的页码
                    unvisited = [p for p in available if p not in visited_pages]
                    
                    if unvisited:
                        # 随机选一个未访问的页码
                        target_page = random.choice(unvisited)
                    else:
                        # 没有未访问的页码，就顺序+1
                        target_page = page_num + 1
                    
                    self.progress(f"   翻页到第 {target_page} 页...")
                    
                    # 方式1（首选）：点击页码按钮
                    clicked = page.evaluate(CLICK_PAGE_JS, target_page)
                    
                    if clicked:
                        self.progress(f"   ✅ 点击页码 {target_page} 成功")
                        page.wait_for_timeout(3000)
                    else:
                        # 方式2：滚动到底部再试
                        self.progress("   点击页码失败，尝试滚动到底部...")
                        try:
                            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                            page.wait_for_timeout(1500)
                            clicked = page.evaluate(CLICK_PAGE_JS, target_page)
                            if clicked:
                                self.progress(f"   ✅ 滚动后点击页码 {target_page} 成功")
                                page.wait_for_timeout(3000)
                            else:
                                # 方式3：尝试点击"下一页"按钮
                                self.progress("   尝试点击'下一页'按钮...")
                                try:
                                    next_btn = page.query_selector('.pagination .next, .page-next, .next-page, .J_Ajax.next, a[class*="next"]')
                                    if next_btn and next_btn.is_visible():
                                        next_btn.click()
                                        page.wait_for_timeout(3000)
                                        self.progress("   ✅ 点击'下一页'成功")
                                    else:
                                        raise Exception("未找到下一页按钮")
                                except:
                                    # 方式4（最后备选）：URL翻页
                                    self.progress(f"   ⚠️ 页码点击均失败，尝试URL翻页")
                                    s = page_num * 44
                                    search_url = "https://s.taobao.com/search?q=" + urllib.parse.quote(KEYWORD) + "&s=" + str(s)
                                    self._navigate_and_handle_login(page, search_url, wait_until="domcontentloaded", timeout=30000)
                                    page.wait_for_timeout(2000)
                        except Exception as e:
                            self.progress(f"   翻页异常: {e}")
                            # 最后备选：URL翻页
                            s = page_num * 44
                            search_url = "https://s.taobao.com/search?q=" + urllib.parse.quote(KEYWORD) + "&s=" + str(s)
                            self._navigate_and_handle_login(page, search_url, wait_until="domcontentloaded", timeout=30000)
                            page.wait_for_timeout(2000)
                
                # 检查滑块验证
                if check_captcha(page):
                    self.progress("⚠️ 搜索页出现滑块验证...")
                    if not handle_captcha(page):
                        if not wait_for_captcha_solve(page):
                            self.progress("❌ 滑块验证未通过，跳过本页")
                            page_num += 1
                            continue
                
                # 确认登录状态
                if not check_login_status(page):
                    self.progress("⚠️ 登录状态失效，重新登录...")
                    login_ok = _wait_for_login(page, ctx, timeout_minutes=5)

                    if not login_ok:
                        self.progress("❌ 重新登录失败")
                        break
                    save_cookies(ctx)
                
                # 滚动页面触发懒加载
                self.progress("   滚动页面加载商品...")
                try:
                    viewport_height = page.evaluate("window.innerHeight")
                    scroll_height = page.evaluate("document.body.scrollHeight")
                    for i in range(3):
                        scroll_to = min(viewport_height * (i + 1) * 0.8, scroll_height - viewport_height)
                        page.evaluate(f"window.scrollTo({{top: {scroll_to}, behavior: 'smooth'}})")
                        page.wait_for_timeout(random.randint(800, 1500))
                    page.evaluate("window.scrollTo({top: 0, behavior: 'smooth'})")
                    page.wait_for_timeout(800)
                except Exception:
                    pass
                
                # 提取商品
                items = page.evaluate(EXTRACT_ITEMS_JS)
                
                if not items:
                    self.progress(f"⚠️ 第{page_num + 1}页无商品数据，可能已到底或被风控")
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    page.wait_for_timeout(2000)
                    items = page.evaluate(EXTRACT_ITEMS_JS)
                    if not items:
                        page_num += 1
                        continue
                
                self.progress(f"   当前页提取到 {len(items)} 个商品")
                
                # 遍历当前页商品
                for item in items:
                    # 检查是否被用户终止
                    if self._stop_event and self._stop_event.is_set():
                        self.progress("🛑 用户终止检测（商品遍历中）")
                        logger.info("用户终止检测（商品遍历中）")
                        break

                    title = item.get('title', '')
                    url = item.get('url', '')
                    if not url or not title:
                        continue
                    
                    total_scanned += 1

                    
                    # 书籍过滤
                    if is_book(title):
                        total_books += 1
                        continue
                    
                    # 商标侵权智能判定（五步核查 + 可选AI增强）
                    location = item.get('location', '')
                    price = item.get('price', '')
                    is_inf, reason, ai_used = is_suspected_infringement(
                        title, price, location,
                        use_ai=True,  # 启用AI增强模式
                        progress_cb=self.progress,
                    )
                    
                    if not is_inf:
                        self.progress(f"   ⏭️ 跳过（{reason}）: {title[:40]}...")
                        continue
                    
                    total_infringing += 1
                    
                    # 侵权商品 -> 进入详情页截图
                    inf_count = len([r for r in results if r["是否侵权"] == "是"])
                    tag = "🤖" if ai_used else "📸"
                    self.progress(f"{tag} 截图侵权商品 #{inf_count + 1}: {title[:50]}...")
                    self.progress(f"   判定依据: {reason}")
                    
                    random_sleep(1.5, 3.5)
                    
                    fpath = ""
                    try:
                        detail_page = ctx.new_page()
                        detail_page.add_init_script(STEALTH_JS)
                        detail_page._progress = self.progress
                        detail_page.goto(url, wait_until="load", timeout=30000)
                        detail_page.wait_for_timeout(1500)
                        
                        if check_captcha(detail_page):
                            self.progress("   ⚠️ 详情页出现滑块验证...")
                            if not handle_captcha(detail_page):
                                wait_for_captcha_solve(detail_page)
                        
                        detail_page.evaluate("window.scrollTo(0, 300)")
                        detail_page.wait_for_timeout(800)
                        
                        fname = f"inf_{inf_count + 1:02d}_{item.get('seller','unknown')[:12]}_{item.get('price','0').replace('.','_')}.png"
                        fpath = str(SCREENSHOTS / fname)
                        detail_page.screenshot(path=fpath, full_page=False)
                        self.progress(f"   ✅ 截图已保存: {fname}")
                        detail_page.close()
                    except Exception as e:
                        self.progress(f"   ❌ 截图失败: {str(e)[:60]}")
                        try: detail_page.close()
                        except: pass
                    
                    results.append({
                        "序号": len(results) + 1,
                        "商品名称": title[:120],
                        "商品URL": url,
                        "记录时间": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "价格": item.get('price', ''),
                        "是否侵权": "是",
                        "截图路径": fpath,
                    })
                
                # 每5页保存一次cookies
                if page_num % 5 == 0 and page_num > 0:
                    save_cookies(ctx)
                
                page_num += 1
            
            # ========== 第三步：生成报告 ==========
            inf_count = len([r for r in results if r["是否侵权"] == "是"])
            
            self.progress("=" * 60)
            self.progress("第三步：生成检测报告")
            self.progress("=" * 60)
            self.progress(f"📊 扫描商品总数: {total_scanned}")
            self.progress(f"🚫 侵权商品: {inf_count}")
            self.progress(f"📚 排除书籍: {total_books}")
            self.progress(f"📍 非北京商家: {total_infringing}")
            self.progress(f"📄 翻页数: {page_num}")
            
            if inf_count == 0:
                self.progress("⚠️ 未找到任何侵权商品")
                ctx.close()
                browser.close()
                return results, inf_count, total_books, total_scanned, ""
            
            self.progress("📝 生成Excel报告...")
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XlImage
            import openpyxl.utils
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = REPORTS_DIR / f"侵权检测报告_{timestamp}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "侵权检测报告"
            
            # 检查是否有AI验证结果，动态决定是否增加AI列
            has_ai_results = any(r.get("ai_result") for r in results)
            
            if has_ai_results:
                headers = ["序号","商品名称","商品截图","商品URL","记录时间","价格","是否侵权","AI验证","AI置信度","AI建议"]
            else:
                headers = ["序号","商品名称","商品截图","商品URL","记录时间","价格","是否侵权"]
            
            hf = Font(bold=True, size=11, color="FFFFFF")
            hb = PatternFill("solid", fgColor="4472C4")
            ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
            
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font=hf; c.fill=hb; c.alignment=ha; c.border=thin
            
            if has_ai_results:
                widths = [8,50,22,55,20,12,12,12,12,18]
            else:
                widths = [8,50,22,55,20,12,12]
            for ci,w in enumerate(widths,1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            
            black = Font(color="000000")
            red = Font(color="FF0000", bold=True)
            green = Font(color="27ae60", bold=True)
            orange = Font(color="e67e22", bold=True)

            for ri, r in enumerate(results, 2):
                c1 = ws.cell(row=ri, column=1, value=r.get("序号", ''))
                c1.border=thin; c1.alignment=Alignment(vertical="center",wrap_text=True); c1.font = black

                c2 = ws.cell(row=ri, column=2, value=r.get("商品名称", ''))
                c2.border=thin; c2.alignment=Alignment(vertical="center",wrap_text=True); c2.font = black

                img_path = r.get("截图路径", "")
                if img_path and os.path.exists(img_path):
                    try:
                        img = XlImage(img_path)
                        img.width = 120
                        img.height = 120 * img.height / img.width if img.width > 0 else 120
                        if img.height > 160: img.height = 160
                        ws.add_image(img, f"C{ri}")
                        ws.row_dimensions[ri].height = max(ws.row_dimensions[ri].height or 0, img.height + 4)
                    except Exception as e:
                        self.progress(f"   嵌入截图失败 (第{ri}行): {e}")

                c4 = ws.cell(row=ri, column=4, value=r.get("商品URL", ''))
                c4.border=thin; c4.alignment=Alignment(vertical="center",wrap_text=True)
                url = r.get("商品URL", '')
                if url:
                    c4.hyperlink = url
                    c4.font = Font(color="0563C1", underline="single")

                c5 = ws.cell(row=ri, column=5, value=r.get("记录时间", ''))
                c5.border=thin; c5.alignment=Alignment(vertical="center",wrap_text=True); c5.font = black

                c6 = ws.cell(row=ri, column=6, value=r.get("价格", ''))
                c6.border=thin; c6.alignment=Alignment(vertical="center",wrap_text=True); c6.font = black

                c7 = ws.cell(row=ri, column=7, value=r.get("是否侵权", ''))
                c7.border=thin; c7.alignment=Alignment(vertical="center",wrap_text=True); c7.font = red

                
                # AI验证结果列（如果有）
                if has_ai_results:
                    ai = r.get("ai_result", {})
                    if ai:
                        ai_inf = ai.get("is_infringement")
                        ai_conf = ai.get("confidence", "")
                        ai_sug = ai.get("suggestion", "")
                        
                        ai_text = "是" if ai_inf is True else ("否" if ai_inf is False else "未知")
                        ai_font = red if ai_inf is True else (green if ai_inf is False else orange)
                        
                        c8 = ws.cell(row=ri, column=8, value=ai_text)
                        c8.border=thin; c8.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True); c8.font = ai_font
                        
                        c9 = ws.cell(row=ri, column=9, value=ai_conf)
                        c9.border=thin; c9.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True); c9.font = ai_font
                        
                        c10 = ws.cell(row=ri, column=10, value=ai_sug)
                        c10.border=thin; c10.alignment=Alignment(vertical="center", wrap_text=True); c10.font = ai_font
                    else:
                        for ci in [8,9,10]:
                            c = ws.cell(row=ri, column=ci, value="-")
                            c.border=thin; c.alignment=Alignment(horizontal="center", vertical="center")

            
            wb.save(str(excel_path))
            self.progress(f"✅ Excel报告已保存: {excel_path}")
            self.progress(f"🖼️  截图保存在: {SCREENSHOTS}")
            
            try: ctx.close()
            except: pass
            try: browser.close()
            except: pass
            
            return results, inf_count, total_books, total_scanned, str(excel_path)


if __name__ == "__main__":
    d = Detector()
    d.run()
